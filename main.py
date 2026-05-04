from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from datetime import datetime
from io import BytesIO
import uuid
import re
import string

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


app = FastAPI(
    title="Scriptora Suite API",
    description="Scriptora Suite: análisis textual, evaluación escritural, proceso, participantes, corpus Excel y exportación.",
    version="0.9.3"
)


# ============================================================
# ALMACENAMIENTO TEMPORAL
# ============================================================

PARTICIPANT_SUBMISSIONS = []


# ============================================================
# MODELOS
# ============================================================

class TextAnalysisRequest(BaseModel):
    text: str
    language: str = "es"
    context: str | None = None
    level: str | None = None
    genre: str | None = None
    purpose: str | None = None


class WritingEvaluationRequest(BaseModel):
    text: str
    language: str = "es"
    level: str | None = None
    genre: str | None = None
    task: str | None = None
    purpose: str | None = None


class WritingProcessRequest(BaseModel):
    text: str
    language: str = "es"
    level: str | None = None
    genre: str | None = None
    task: str | None = None
    purpose: str | None = None

    writing_mode: str | None = "live"
    total_time_seconds: float | None = 0
    initial_latency_seconds: float | None = 0
    pause_count: int | None = 0
    long_pause_count: int | None = 0
    edit_count: int | None = 0
    deletion_count: int | None = 0
    insertion_count: int | None = 0
    local_adjustment_count: int | None = 0
    reformulation_count: int | None = 0
    expansion_count: int | None = 0
    reduction_count: int | None = 0
    macro_adjustment_count: int | None = 0
    max_text_length: int | None = 0
    final_text_length: int | None = 0
    input_event_count: int | None = 0


class ExcelExportRequest(WritingProcessRequest):
    analysis_mode: str = "text_analysis"
    entry_mode: str = "single_text"


class MultiTextRequest(BaseModel):
    raw_input: str
    analysis_mode: str = "text_analysis"
    language: str = "es"
    level: str | None = "general"
    genre: str | None = "general"
    purpose: str | None = "multitext_analysis"


class MultiExcelExportRequest(MultiTextRequest):
    pass


class ParticipantSubmissionRequest(BaseModel):
    participant_id: str | None = None
    text: str
    language: str = "es"
    level: str | None = "general"
    genre: str | None = "argumentativo"
    task: str | None = "participant_writing_task"
    purpose: str | None = "silent_participant_submission"


# ============================================================
# FUNCIONES BÁSICAS
# ============================================================

def split_sentences(text: str):
    sentences = re.split(r'[.!?¿¡]+', text)
    return [s.strip() for s in sentences if s.strip()]


def tokenize_words(text: str):
    text = text.lower()
    text = text.translate(str.maketrans('', '', string.punctuation + "¿¡“”‘’"))
    return [w for w in text.split() if w.strip()]


def estimate_paragraphs(text: str):
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
    return len(paragraphs)


def estimate_lexical_density(words):
    if not words:
        return 0

    function_words = {
        "el", "la", "los", "las", "un", "una", "unos", "unas",
        "de", "del", "a", "al", "en", "con", "por", "para", "sin",
        "y", "o", "pero", "que", "se", "me", "te", "lo", "le", "les",
        "es", "son", "fue", "era", "ser", "estar", "está", "están",
        "como", "más", "menos", "muy", "también", "no", "sí",
        "su", "sus", "mi", "mis", "tu", "tus", "este", "esta",
        "estos", "estas", "ese", "esa", "eso", "hay", "ha", "han",
        "he", "hemos", "fui", "fueron", "tiene", "tienen", "tengo",
        "cuando", "donde", "quien", "quienes", "cual", "cuales"
    }

    content_words = [w for w in words if w not in function_words]
    return round(len(content_words) / len(words), 3)


def count_connectors(text: str):
    connectors = [
        "porque", "por lo tanto", "sin embargo", "además", "también",
        "aunque", "en cambio", "por ejemplo", "finalmente", "primero",
        "segundo", "luego", "después", "entonces", "así", "ya que",
        "debido a", "en conclusión", "por otra parte", "por consiguiente",
        "a pesar de", "en primer lugar", "en segundo lugar", "por ende",
        "de este modo", "en síntesis"
    ]

    text_lower = text.lower()
    count = 0
    found = []

    for connector in connectors:
        occurrences = text_lower.count(connector)
        if occurrences > 0:
            count += occurrences
            found.append(connector)

    return count, sorted(list(set(found)))


def count_punctuation_marks(text: str):
    marks = [".", ",", ";", ":", "¿", "?", "¡", "!"]
    return sum(text.count(m) for m in marks)


def detect_closure(text: str):
    closure_markers = [
        "en conclusión",
        "finalmente",
        "en síntesis",
        "por último",
        "para concluir",
        "en resumen"
    ]
    lower_text = text.lower()
    return any(marker in lower_text for marker in closure_markers)


# ============================================================
# SCRIPTORA T
# ============================================================

def interpret_text_metrics(word_count, avg_sentence_length, lexical_density, ttr):
    comments = []

    if word_count < 50:
        comments.append("El texto es breve; las métricas deben interpretarse con cautela.")
    elif word_count <= 250:
        comments.append("El texto posee una extensión suficiente para una caracterización inicial.")
    else:
        comments.append("El texto posee una extensión amplia, adecuada para un análisis más estable.")

    if avg_sentence_length < 12:
        comments.append("La longitud oracional promedio es baja, compatible con estructura sintáctica simple.")
    elif avg_sentence_length <= 22:
        comments.append("La longitud oracional promedio se ubica en un rango medio.")
    else:
        comments.append("La longitud oracional promedio es alta, lo que puede incrementar la complejidad de procesamiento.")

    if lexical_density < 0.45:
        comments.append("La densidad léxica estimada es baja.")
    elif lexical_density <= 0.65:
        comments.append("La densidad léxica estimada es media.")
    else:
        comments.append("La densidad léxica estimada es alta, compatible con mayor concentración informativa.")

    if ttr < 0.35:
        comments.append("La diversidad léxica superficial es baja.")
    elif ttr <= 0.60:
        comments.append("La diversidad léxica superficial se encuentra en un rango medio.")
    else:
        comments.append("La diversidad léxica superficial es alta, aunque en textos breves este valor puede sobreestimarse.")

    return " ".join(comments)


def analyze_text_core(text: str):
    text = text.strip()
    words = tokenize_words(text)
    sentences = split_sentences(text)

    word_count = len(words)
    char_count = len(text)
    sentence_count = len(sentences)
    unique_words = len(set(words))
    avg_sentence_length = round(word_count / sentence_count, 2) if sentence_count > 0 else 0
    ttr = round(unique_words / word_count, 3) if word_count > 0 else 0
    lexical_density = estimate_lexical_density(words)

    interpretation = interpret_text_metrics(
        word_count,
        avg_sentence_length,
        lexical_density,
        ttr
    )

    return {
        "word_count": word_count,
        "char_count": char_count,
        "sentence_count": sentence_count,
        "unique_words": unique_words,
        "avg_sentence_length": avg_sentence_length,
        "type_token_ratio": ttr,
        "lexical_density_proxy": lexical_density,
        "interpretation_text": interpretation
    }


# ============================================================
# SCRIPTORA W PRODUCTO
# ============================================================

def score_writing_product(text, words=None, sentences=None, genre="general"):
    text = text.strip()
    words = words if words is not None else tokenize_words(text)
    sentences = sentences if sentences is not None else split_sentences(text)

    word_count = len(words)
    sentence_count = len(sentences)
    paragraph_count = estimate_paragraphs(text)
    connector_count, connectors_found = count_connectors(text)
    punctuation_count = count_punctuation_marks(text)
    lexical_density = estimate_lexical_density(words)

    avg_sentence_length = round(word_count / sentence_count, 2) if sentence_count > 0 else 0
    unique_words = len(set(words))
    ttr = round(unique_words / word_count, 3) if word_count > 0 else 0

    extension_score = min(100, round((word_count / 250) * 100)) if word_count > 0 else 0

    if paragraph_count >= 3:
        organization_score = 85
    elif paragraph_count == 2:
        organization_score = 70
    elif paragraph_count == 1 and word_count > 80:
        organization_score = 55
    else:
        organization_score = 40

    if connector_count >= 6:
        cohesion_score = 90
    elif connector_count >= 3:
        cohesion_score = 75
    elif connector_count >= 1:
        cohesion_score = 60
    else:
        cohesion_score = 40

    if 10 <= avg_sentence_length <= 24:
        syntax_score = 80
    elif avg_sentence_length < 10:
        syntax_score = 60
    else:
        syntax_score = 65

    if ttr >= 0.55:
        lexical_variety_score = 85
    elif ttr >= 0.40:
        lexical_variety_score = 70
    else:
        lexical_variety_score = 55

    if lexical_density >= 0.65:
        informational_density_score = 85
    elif lexical_density >= 0.45:
        informational_density_score = 70
    else:
        informational_density_score = 55

    if sentence_count >= 3 and word_count >= 80:
        elaboration_score = 80
    elif sentence_count >= 2 and word_count >= 40:
        elaboration_score = 65
    elif word_count >= 20:
        elaboration_score = 50
    else:
        elaboration_score = 35

    if connector_count >= 2 and paragraph_count >= 2:
        global_coherence_score = 80
    elif connector_count >= 1 or paragraph_count >= 2:
        global_coherence_score = 65
    else:
        global_coherence_score = 45

    if punctuation_count >= 6:
        punctuation_control_score = 80
    elif punctuation_count >= 3:
        punctuation_control_score = 65
    elif punctuation_count >= 1:
        punctuation_control_score = 50
    else:
        punctuation_control_score = 35

    closure_present = detect_closure(text)
    closure_score = 80 if closure_present else 45

    lower_text = text.lower()
    genre_score = 65
    genre_comment = "La adecuación al género se estima de manera preliminar."

    if genre == "argumentativo":
        argument_markers = [
            "creo", "pienso", "opino", "debería", "por lo tanto",
            "en conclusión", "argumento", "razón", "postura",
            "estoy de acuerdo", "no estoy de acuerdo", "a favor", "en contra"
        ]
        found_argument_markers = sum(1 for m in argument_markers if m in lower_text)
        genre_score = min(95, 55 + found_argument_markers * 8)
        genre_comment = "En el texto argumentativo se observaron marcas preliminares de postura, justificación o cierre."

    elif genre == "narrativo":
        narrative_markers = [
            "un día", "luego", "después", "entonces", "finalmente",
            "personaje", "historia", "había", "ocurrió", "mientras"
        ]
        found_narrative_markers = sum(1 for m in narrative_markers if m in lower_text)
        genre_score = min(95, 55 + found_narrative_markers * 8)
        genre_comment = "En el texto narrativo se observaron marcas preliminares de secuencia o progresión temporal."

    elif genre == "expositivo":
        expository_markers = [
            "es decir", "por ejemplo", "se define", "consiste",
            "presenta", "explica", "corresponde", "se caracteriza"
        ]
        found_expository_markers = sum(1 for m in expository_markers if m in lower_text)
        genre_score = min(95, 55 + found_expository_markers * 8)
        genre_comment = "En el texto expositivo se observaron marcas preliminares de explicación o desarrollo informativo."

    global_score = round(
        extension_score * 0.10 +
        organization_score * 0.12 +
        cohesion_score * 0.12 +
        syntax_score * 0.10 +
        lexical_variety_score * 0.10 +
        informational_density_score * 0.10 +
        elaboration_score * 0.14 +
        global_coherence_score * 0.12 +
        genre_score * 0.07 +
        closure_score * 0.03,
        1
    )

    if global_score < 50:
        level_label = "Inicial"
    elif global_score < 65:
        level_label = "En desarrollo"
    elif global_score < 80:
        level_label = "Adecuado"
    else:
        level_label = "Avanzado"

    interpretation = (
        f"El desempeño escritural preliminar se ubica en nivel {level_label}. "
        f"La evaluación considera extensión, organización, cohesión, sintaxis, diversidad léxica, "
        f"densidad informativa, elaboración de ideas, coherencia global, adecuación al género y cierre textual. "
        f"{genre_comment} "
        f"Esta estimación es exploratoria y deberá calibrarse con rúbricas humanas, benchmarks y datos TRUNAJOD."
    )

    return {
        "word_count": word_count,
        "sentence_count": sentence_count,
        "paragraph_count": paragraph_count,
        "avg_sentence_length": avg_sentence_length,
        "unique_words": unique_words,
        "type_token_ratio": ttr,
        "lexical_density_proxy": lexical_density,
        "connector_count": connector_count,
        "connectors_found": connectors_found,
        "punctuation_count": punctuation_count,
        "closure_present": closure_present,
        "scores": {
            "extension": extension_score,
            "organization": organization_score,
            "cohesion": cohesion_score,
            "syntax_control": syntax_score,
            "lexical_variety": lexical_variety_score,
            "informational_density": informational_density_score,
            "idea_elaboration": elaboration_score,
            "global_coherence": global_coherence_score,
            "punctuation_control": punctuation_control_score,
            "genre_adequacy": genre_score,
            "textual_closure": closure_score,
            "global_writing_score": global_score
        },
        "level_label": level_label,
        "interpretation": interpretation
    }


# ============================================================
# SCRIPTORA W PROCESO
# ============================================================

def score_writing_process(request: WritingProcessRequest):
    text = request.text.strip()

    writing_mode = request.writing_mode or "live"
    total_time = request.total_time_seconds or 0
    initial_latency = request.initial_latency_seconds or 0
    long_pause_count = request.long_pause_count or 0
    edit_count = request.edit_count or 0
    deletion_count = request.deletion_count or 0
    insertion_count = request.insertion_count or 0
    local_adjustment_count = request.local_adjustment_count or 0
    reformulation_count = request.reformulation_count or 0
    expansion_count = request.expansion_count or 0
    reduction_count = request.reduction_count or 0
    macro_adjustment_count = request.macro_adjustment_count or 0
    max_text_length = request.max_text_length or len(text)
    final_text_length = request.final_text_length or len(text)
    input_event_count = request.input_event_count or 0

    words = tokenize_words(text)
    word_count = len(words)

    words_per_minute = round((word_count / total_time) * 60, 2) if total_time > 0 else 0
    final_stability_ratio = round(final_text_length / max_text_length, 3) if max_text_length > 0 else 1

    if writing_mode == "pasted":
        return {
            "writing_mode": writing_mode,
            "total_time_seconds": total_time,
            "initial_latency_seconds": initial_latency,
            "long_pause_count": long_pause_count,
            "edit_count": edit_count,
            "deletion_count": deletion_count,
            "insertion_count": insertion_count,
            "local_adjustment_count": local_adjustment_count,
            "reformulation_count": reformulation_count,
            "expansion_count": expansion_count,
            "reduction_count": reduction_count,
            "macro_adjustment_count": macro_adjustment_count,
            "max_text_length": max_text_length,
            "final_text_length": final_text_length,
            "input_event_count": input_event_count,
            "words_per_minute": words_per_minute,
            "final_stability_ratio": final_stability_ratio,
            "planning_score": 0,
            "monitoring_score": 0,
            "revision_score": 0,
            "reformulation_score": 0,
            "fluency_score": 0,
            "recursivity_score": 0,
            "process_regulation_score": 0,
            "process_regulation_label": "No trazable",
            "interpretation": (
                "El texto parece haber sido pegado o ingresado sin trazabilidad suficiente del proceso. "
                "La evaluación del producto textual puede realizarse, pero la regulación escritural no debe inferirse desde estos datos."
            )
        }

    planning_score = 80 if initial_latency >= 5 else 60 if initial_latency >= 2 else 40
    monitoring_score = 85 if long_pause_count >= 3 else 65 if long_pause_count >= 1 else 35

    revision_events = deletion_count + local_adjustment_count + reduction_count
    revision_score = 85 if revision_events >= 12 else 70 if revision_events >= 5 else 55 if revision_events >= 1 else 35

    reformulation_score = 85 if reformulation_count >= 3 else 70 if reformulation_count >= 1 else 35

    if 8 <= words_per_minute <= 40:
        fluency_score = 80
    elif words_per_minute > 40:
        fluency_score = 60
    elif words_per_minute > 0:
        fluency_score = 55
    else:
        fluency_score = 35

    recursivity_score = 85 if macro_adjustment_count >= 2 else 70 if macro_adjustment_count == 1 else 45

    process_score = round(
        planning_score * 0.15 +
        monitoring_score * 0.20 +
        revision_score * 0.20 +
        reformulation_score * 0.15 +
        fluency_score * 0.15 +
        recursivity_score * 0.15,
        1
    )

    if process_score < 45:
        process_label = "Regulación baja"
    elif process_score < 65:
        process_label = "Regulación emergente"
    elif process_score < 80:
        process_label = "Regulación media"
    else:
        process_label = "Regulación alta"

    parts = []

    if initial_latency < 2:
        parts.append("La latencia inicial fue baja, lo que sugiere inicio rápido de la escritura.")
    else:
        parts.append("La latencia inicial sugiere una fase inicial de planificación.")

    if long_pause_count == 0:
        parts.append("Se observaron pocas pausas largas, con baja evidencia de monitoreo reflexivo.")
    else:
        parts.append("Las pausas largas sugieren momentos de monitoreo, planificación o revisión.")

    if local_adjustment_count > 0:
        parts.append("Se detectaron ajustes locales compatibles con control microtextual.")

    if reformulation_count > 0:
        parts.append("Se detectaron reformulaciones, lo que sugiere revisión de segmentos mayores.")

    if macro_adjustment_count > 0:
        parts.append("Se observaron ajustes macrotextuales compatibles con reorganización del texto.")

    interpretation = (
        f"La regulación escritural preliminar se ubica en: {process_label}. "
        + " ".join(parts)
        + " Esta estimación es exploratoria y deberá calibrarse con datos reales de escritura."
    )

    return {
        "writing_mode": writing_mode,
        "total_time_seconds": total_time,
        "initial_latency_seconds": initial_latency,
        "long_pause_count": long_pause_count,
        "edit_count": edit_count,
        "deletion_count": deletion_count,
        "insertion_count": insertion_count,
        "local_adjustment_count": local_adjustment_count,
        "reformulation_count": reformulation_count,
        "expansion_count": expansion_count,
        "reduction_count": reduction_count,
        "macro_adjustment_count": macro_adjustment_count,
        "max_text_length": max_text_length,
        "final_text_length": final_text_length,
        "input_event_count": input_event_count,
        "words_per_minute": words_per_minute,
        "final_stability_ratio": final_stability_ratio,
        "planning_score": planning_score,
        "monitoring_score": monitoring_score,
        "revision_score": revision_score,
        "reformulation_score": reformulation_score,
        "fluency_score": fluency_score,
        "recursivity_score": recursivity_score,
        "process_regulation_score": process_score,
        "process_regulation_label": process_label,
        "interpretation": interpretation
    }


def integrated_writing_interpretation(product_metrics, process_metrics):
    product_label = product_metrics.get("level_label", "No determinado")
    process_label = process_metrics.get("process_regulation_label", "No determinado")
    product_score = product_metrics.get("scores", {}).get("global_writing_score", 0)
    process_score = process_metrics.get("process_regulation_score", 0)

    if process_label == "No trazable":
        return (
            f"El producto escrito se ubica preliminarmente en nivel {product_label}. "
            "Sin embargo, no existe trazabilidad suficiente para interpretar el proceso de escritura. "
            "La síntesis integrada queda limitada al análisis del producto textual."
        )

    if product_score < 50 and process_score < 50:
        return (
            "La síntesis integrada sugiere bajo desarrollo del producto y baja evidencia de regulación escritural. "
            "Esto puede corresponder a una producción breve, poco elaborada y con escasa revisión."
        )

    if product_score < 50 and process_score >= 65:
        return (
            "La síntesis integrada muestra una tensión relevante: el producto escrito aún es bajo, "
            "pero el proceso evidencia señales de planificación, monitoreo o revisión. "
            "Esto puede indicar esfuerzo regulatorio que todavía no logra consolidarse en el texto final."
        )

    if product_score >= 65 and process_score < 50:
        return (
            "La síntesis integrada sugiere un producto relativamente adecuado, pero con pocas señales de regulación observable. "
            "Esto podría corresponder a escritura fluida, automatizada o a una captura insuficiente del proceso."
        )

    return (
        f"La síntesis integrada muestra un producto en nivel {product_label} junto con un proceso clasificado como {process_label}. "
        "El resultado sugiere una relación positiva entre elaboración textual y control regulatorio durante la escritura."
    )


# ============================================================
# MULTITEXTO Y CORPUS
# ============================================================

def parse_multitext_input(raw_input: str):
    raw_input = raw_input.strip()
    if not raw_input:
        return []

    pattern = r"###\s*ID:\s*(.+)"
    matches = list(re.finditer(pattern, raw_input))
    records = []

    if not matches:
        blocks = [b.strip() for b in raw_input.split("\n---\n") if b.strip()]
        for idx, block in enumerate(blocks, start=1):
            records.append({
                "id": f"texto_{idx:03d}",
                "text": block
            })
        return records

    for i, match in enumerate(matches):
        current_id = match.group(1).strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(raw_input)
        text_block = raw_input[start:end].strip()

        if text_block:
            records.append({
                "id": current_id,
                "text": text_block
            })

    return records


def flatten_product_record(source, product_metrics):
    record = dict(source)

    record.update({
        "word_count": product_metrics.get("word_count"),
        "sentence_count": product_metrics.get("sentence_count"),
        "paragraph_count": product_metrics.get("paragraph_count"),
        "avg_sentence_length": product_metrics.get("avg_sentence_length"),
        "unique_words": product_metrics.get("unique_words"),
        "type_token_ratio": product_metrics.get("type_token_ratio"),
        "lexical_density_proxy": product_metrics.get("lexical_density_proxy"),
        "connector_count": product_metrics.get("connector_count"),
        "connectors_found": ", ".join(product_metrics.get("connectors_found", [])),
        "punctuation_count": product_metrics.get("punctuation_count"),
        "closure_present": product_metrics.get("closure_present"),
        "scores_extension": product_metrics.get("scores", {}).get("extension"),
        "scores_organization": product_metrics.get("scores", {}).get("organization"),
        "scores_cohesion": product_metrics.get("scores", {}).get("cohesion"),
        "scores_syntax_control": product_metrics.get("scores", {}).get("syntax_control"),
        "scores_lexical_variety": product_metrics.get("scores", {}).get("lexical_variety"),
        "scores_informational_density": product_metrics.get("scores", {}).get("informational_density"),
        "scores_idea_elaboration": product_metrics.get("scores", {}).get("idea_elaboration"),
        "scores_global_coherence": product_metrics.get("scores", {}).get("global_coherence"),
        "scores_punctuation_control": product_metrics.get("scores", {}).get("punctuation_control"),
        "scores_genre_adequacy": product_metrics.get("scores", {}).get("genre_adequacy"),
        "scores_textual_closure": product_metrics.get("scores", {}).get("textual_closure"),
        "scores_global_writing_score": product_metrics.get("scores", {}).get("global_writing_score"),
        "level_label": product_metrics.get("level_label"),
        "interpretation_product": product_metrics.get("interpretation")
    })

    return record


def flatten_text_record(source, text_metrics):
    record = dict(source)

    record.update({
        "word_count": text_metrics.get("word_count"),
        "char_count": text_metrics.get("char_count"),
        "sentence_count": text_metrics.get("sentence_count"),
        "unique_words": text_metrics.get("unique_words"),
        "avg_sentence_length": text_metrics.get("avg_sentence_length"),
        "type_token_ratio": text_metrics.get("type_token_ratio"),
        "lexical_density_proxy": text_metrics.get("lexical_density_proxy"),
        "interpretation_text": text_metrics.get("interpretation_text")
    })

    return record


def analyze_multitext(request: MultiTextRequest):
    records = parse_multitext_input(request.raw_input)
    results = []

    for record in records:
        item_id = record["id"]
        text = record["text"]

        source = {
            "id": item_id,
            "analysis_mode": request.analysis_mode,
            "language": request.language,
            "level": request.level,
            "genre": request.genre,
            "texto_original": text
        }

        if request.analysis_mode == "text_analysis":
            metrics = analyze_text_core(text)
            output_row = flatten_text_record(source, metrics)
            output_row["module"] = "Scriptora T · Multitexto"
        else:
            product = score_writing_product(text=text, genre=request.genre or "general")
            output_row = flatten_product_record(source, product)
            output_row["module"] = "Scriptora W · Producto multitexto"

        results.append(output_row)

    return results


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def process_corpus_excel(file_bytes: bytes, analysis_mode: str = "writing_product"):
    wb_in = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb_in.active

    headers = []
    for cell in ws[1]:
        headers.append(normalize_header(cell.value))

    if "texto" not in headers and "text" not in headers:
        raise ValueError("El Excel debe incluir una columna llamada 'texto' o 'text'.")

    text_col = "texto" if "texto" in headers else "text"
    text_idx = headers.index(text_col)

    rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        original = {}
        for h, v in zip(headers, row):
            if h:
                original[h] = v if v is not None else ""

        text = str(row[text_idx]).strip() if row[text_idx] is not None else ""

        if not text:
            continue

        item_id = original.get("id") or original.get("participant_id") or f"texto_{row_idx:03d}"
        genre = original.get("genero") or original.get("genre") or "general"
        level = original.get("nivel") or original.get("level") or "general"

        source = dict(original)
        source["id"] = item_id
        source["texto_original"] = text
        source["analysis_mode"] = analysis_mode
        source["genre_used"] = genre
        source["level_used"] = level

        if analysis_mode == "text_analysis":
            metrics = analyze_text_core(text)
            output_row = flatten_text_record(source, metrics)
        else:
            product_metrics = score_writing_product(text=text, genre=str(genre).lower())
            output_row = flatten_product_record(source, product_metrics)

        rows.append(output_row)

    return rows


# ============================================================
# PARTICIPANTES
# ============================================================

def create_participant_record(request: ParticipantSubmissionRequest):
    submission_id = str(uuid.uuid4())
    participant_id = request.participant_id.strip() if request.participant_id else f"participante_{len(PARTICIPANT_SUBMISSIONS) + 1:03d}"
    timestamp = datetime.utcnow().isoformat()
    text = request.text.strip()

    product_metrics = score_writing_product(
        text=text,
        genre=request.genre or "argumentativo"
    )

    record = flatten_product_record(
        source={
            "submission_id": submission_id,
            "participant_id": participant_id,
            "timestamp_utc": timestamp,
            "language": request.language,
            "level": request.level,
            "genre": request.genre,
            "task": request.task,
            "purpose": request.purpose,
            "texto_original": text
        },
        product_metrics=product_metrics
    )

    PARTICIPANT_SUBMISSIONS.append(record)
    return record


# ============================================================
# EXCEL
# ============================================================

def safe_value(value):
    if isinstance(value, (dict, list)):
        return str(value)
    if value is None:
        return ""
    return value


def flatten_dict(data, prefix=""):
    flat = {}

    for key, value in data.items():
        new_key = f"{prefix}_{key}" if prefix else key

        if isinstance(value, dict):
            flat.update(flatten_dict(value, new_key))
        elif isinstance(value, list):
            flat[new_key] = ", ".join([str(v) for v in value])
        else:
            flat[new_key] = value

    return flat


def build_variable_dictionary():
    return [
        {
            "variable": "id",
            "nombre_amigable": "ID del texto",
            "modulo": "Corpus / Multitexto",
            "dimension": "Identificación",
            "descripcion": "Identificador del texto, sujeto o registro analizado.",
            "como_se_calcula": "Asignado por el usuario o generado automáticamente.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite vincular resultados con sujetos, textos o registros.",
            "observaciones": "Clave para análisis por corpus."
        },
        {
            "variable": "participant_id",
            "nombre_amigable": "ID del participante",
            "modulo": "Participante",
            "dimension": "Identificación",
            "descripcion": "Identificador asignado al participante.",
            "como_se_calcula": "Ingresado por usuario o generado automáticamente.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite vincular respuestas con participantes.",
            "observaciones": "En investigación real debe definirse anonimización."
        },
        {
            "variable": "submission_id",
            "nombre_amigable": "ID del envío",
            "modulo": "Participante",
            "dimension": "Identificación",
            "descripcion": "Identificador único de la respuesta enviada.",
            "como_se_calcula": "UUID generado automáticamente.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite rastrear cada envío.",
            "observaciones": "Útil cuando un participante responde más de una tarea."
        },
        {
            "variable": "analysis_id",
            "nombre_amigable": "ID del análisis",
            "modulo": "Metadatos",
            "dimension": "Identificación",
            "descripcion": "Identificador único del análisis realizado.",
            "como_se_calcula": "UUID generado automáticamente.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite rastrear cada análisis.",
            "observaciones": "Útil para bases con múltiples sujetos."
        },
        {
            "variable": "analysis_mode",
            "nombre_amigable": "Línea de análisis",
            "modulo": "Configuración",
            "dimension": "Diseño analítico",
            "descripcion": "Indica si el análisis corresponde a texto descriptivo, producto escrito o proceso escritural.",
            "como_se_calcula": "Seleccionado por el investigador.",
            "tipo_valor": "Texto",
            "rango_esperado": "text_analysis / writing_product / writing_process",
            "interpretacion_general": "Distingue análisis textual de evaluación escritural.",
            "observaciones": "Clave para evitar mezclar texto y escritura."
        },
        {
            "variable": "entry_mode",
            "nombre_amigable": "Tipo de entrada",
            "modulo": "Configuración",
            "dimension": "Ingreso de datos",
            "descripcion": "Indica si el dato fue ingresado como texto individual, multitexto, corpus Excel o proceso.",
            "como_se_calcula": "Seleccionado por el investigador.",
            "tipo_valor": "Texto",
            "rango_esperado": "single_text / pasted_multi / corpus_excel / live_process",
            "interpretacion_general": "Permite identificar la vía de ingreso.",
            "observaciones": "No todos los modos aplican a todas las líneas de análisis."
        },
        {
            "variable": "timestamp_utc",
            "nombre_amigable": "Fecha y hora UTC",
            "modulo": "Metadatos",
            "dimension": "Trazabilidad",
            "descripcion": "Fecha y hora de generación del análisis.",
            "como_se_calcula": "datetime.utcnow().isoformat()",
            "tipo_valor": "Fecha/hora",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite ordenar cronológicamente los análisis.",
            "observaciones": "Está en horario UTC."
        },
        {
            "variable": "texto_original",
            "nombre_amigable": "Texto original",
            "modulo": "Todos",
            "dimension": "Entrada",
            "descripcion": "Texto ingresado por el usuario o cargado desde corpus.",
            "como_se_calcula": "Se conserva el texto enviado.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Permite revisar manualmente el insumo.",
            "observaciones": "Debe manejarse con cuidado si contiene información sensible."
        },
        {
            "variable": "word_count",
            "nombre_amigable": "Número de palabras",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Extensión",
            "descripcion": "Cantidad total de palabras detectadas en el texto.",
            "como_se_calcula": "Conteo de tokens después de limpieza básica.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Mayor extensión puede sugerir mayor desarrollo textual, aunque no implica necesariamente mejor calidad.",
            "observaciones": "Debe interpretarse según nivel, tarea y género."
        },
        {
            "variable": "char_count",
            "nombre_amigable": "Número de caracteres",
            "modulo": "Scriptora T",
            "dimension": "Extensión",
            "descripcion": "Cantidad total de caracteres del texto original.",
            "como_se_calcula": "Conteo directo de caracteres del texto.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Permite estimar longitud bruta del texto.",
            "observaciones": "Incluye espacios y signos."
        },
        {
            "variable": "sentence_count",
            "nombre_amigable": "Número de oraciones",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Estructura textual",
            "descripcion": "Cantidad de unidades oracionales detectadas.",
            "como_se_calcula": "Segmentación por signos de cierre.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Permite estimar organización básica y longitud oracional promedio.",
            "observaciones": "La segmentación es preliminar."
        },
        {
            "variable": "paragraph_count",
            "nombre_amigable": "Número de párrafos",
            "modulo": "Scriptora W Producto",
            "dimension": "Organización",
            "descripcion": "Cantidad de párrafos separados por saltos de línea.",
            "como_se_calcula": "Conteo de bloques no vacíos separados por salto de línea.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Más de un párrafo puede reflejar mayor organización discursiva.",
            "observaciones": "En textos breves puede no esperarse más de un párrafo."
        },
        {
            "variable": "avg_sentence_length",
            "nombre_amigable": "Longitud oracional promedio",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Sintaxis",
            "descripcion": "Promedio de palabras por oración.",
            "como_se_calcula": "word_count / sentence_count",
            "tipo_valor": "Decimal",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Valores bajos sugieren estructuras simples; valores altos pueden aumentar complejidad.",
            "observaciones": "Debe interpretarse según edad, género y tarea."
        },
        {
            "variable": "unique_words",
            "nombre_amigable": "Palabras únicas",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Diversidad léxica",
            "descripcion": "Cantidad de formas léxicas distintas.",
            "como_se_calcula": "Conteo de tokens únicos.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Aporta información inicial sobre variedad léxica.",
            "observaciones": "Depende de la extensión."
        },
        {
            "variable": "type_token_ratio",
            "nombre_amigable": "TTR",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Diversidad léxica",
            "descripcion": "Proporción entre palabras únicas y total de palabras.",
            "como_se_calcula": "unique_words / word_count",
            "tipo_valor": "Decimal",
            "rango_esperado": "0 a 1",
            "interpretacion_general": "Valores más altos sugieren mayor diversidad léxica superficial.",
            "observaciones": "En textos breves puede sobreestimar diversidad."
        },
        {
            "variable": "lexical_density_proxy",
            "nombre_amigable": "Densidad léxica estimada",
            "modulo": "Scriptora T / Scriptora W Producto",
            "dimension": "Densidad informativa",
            "descripcion": "Proporción estimada de palabras de contenido.",
            "como_se_calcula": "Palabras no funcionales / total de palabras.",
            "tipo_valor": "Decimal",
            "rango_esperado": "0 a 1",
            "interpretacion_general": "Valores altos sugieren mayor concentración informativa.",
            "observaciones": "Aproximación inicial."
        },
        {
            "variable": "connector_count",
            "nombre_amigable": "Conectores detectados",
            "modulo": "Scriptora W Producto",
            "dimension": "Cohesión",
            "descripcion": "Cantidad de conectores discursivos reconocidos.",
            "como_se_calcula": "Búsqueda en lista preliminar de conectores.",
            "tipo_valor": "Entero",
            "rango_esperado": "0 en adelante",
            "interpretacion_general": "Más conectores pueden indicar mayor articulación discursiva.",
            "observaciones": "No evalúa pertinencia semántica."
        },
        {
            "variable": "scores_global_writing_score",
            "nombre_amigable": "Puntaje global del producto escrito",
            "modulo": "Scriptora W Producto",
            "dimension": "Desempeño escritural",
            "descripcion": "Índice sintético de calidad preliminar del texto final.",
            "como_se_calcula": "Combinación ponderada de variables de producto.",
            "tipo_valor": "Decimal",
            "rango_esperado": "0 a 100",
            "interpretacion_general": "Valores más altos sugieren mejor desempeño escritural preliminar.",
            "observaciones": "Debe calibrarse con rúbricas humanas."
        },
        {
            "variable": "level_label",
            "nombre_amigable": "Nivel del producto",
            "modulo": "Scriptora W Producto",
            "dimension": "Desempeño escritural",
            "descripcion": "Categoría interpretativa del producto escrito.",
            "como_se_calcula": "Clasificación del puntaje global en rangos.",
            "tipo_valor": "Texto",
            "rango_esperado": "Inicial / En desarrollo / Adecuado / Avanzado",
            "interpretacion_general": "Resume el desempeño escritural preliminar.",
            "observaciones": "Los puntos de corte son prototípicos."
        },
        {
            "variable": "interpretation_text",
            "nombre_amigable": "Interpretación textual",
            "modulo": "Scriptora T",
            "dimension": "Interpretación",
            "descripcion": "Comentario interpretativo descriptivo sobre el texto.",
            "como_se_calcula": "Reglas interpretativas basadas en métricas textuales.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Sintetiza características textuales.",
            "observaciones": "No equivale a evaluación de escritura."
        },
        {
            "variable": "interpretation_product",
            "nombre_amigable": "Interpretación del producto escrito",
            "modulo": "Scriptora W Producto",
            "dimension": "Interpretación",
            "descripcion": "Comentario interpretativo preliminar sobre el producto escrito.",
            "como_se_calcula": "Reglas interpretativas basadas en métricas y puntajes.",
            "tipo_valor": "Texto",
            "rango_esperado": "Variable",
            "interpretacion_general": "Sintetiza el resultado para lectura cualitativa.",
            "observaciones": "Requiere validación empírica."
        }
    ]


def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value_length = len(str(cell.value)) if cell.value is not None else 0
            max_length = max(max_length, value_length)

        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def get_dictionary_info(variable):
    clean_variable = variable

    if clean_variable.startswith("product_"):
        clean_variable = clean_variable.replace("product_", "", 1)

    if clean_variable.startswith("process_"):
        clean_variable = clean_variable.replace("process_", "", 1)

    dictionary_map = {row["variable"]: row for row in build_variable_dictionary()}
    return dictionary_map.get(clean_variable, {})


def add_dictionary_sheet(wb):
    ws_dict = wb.create_sheet("diccionario_variables")
    dictionary_rows = build_variable_dictionary()

    dict_headers = [
        "variable",
        "nombre_amigable",
        "modulo",
        "dimension",
        "descripcion",
        "como_se_calcula",
        "tipo_valor",
        "rango_esperado",
        "interpretacion_general",
        "observaciones"
    ]

    ws_dict.append(dict_headers)

    for row in dictionary_rows:
        ws_dict.append([row.get(h, "") for h in dict_headers])

    style_sheet(ws_dict)


def add_metadata_sheet(wb, metadata):
    ws_meta = wb.create_sheet("metadatos")
    ws_meta.append(["campo", "valor"])

    for key, value in metadata.items():
        ws_meta.append([key, safe_value(value)])

    style_sheet(ws_meta)


def build_results_vertical_rows(flat_results):
    rows = []

    for variable, value in flat_results.items():
        dict_info = get_dictionary_info(variable)

        rows.append({
            "variable": variable,
            "valor": safe_value(value),
            "nombre_amigable": dict_info.get("nombre_amigable", ""),
            "modulo": dict_info.get("modulo", ""),
            "dimension": dict_info.get("dimension", ""),
            "descripcion": dict_info.get("descripcion", "")
        })

    return rows


def create_single_analysis_excel(export_data):
    wb = Workbook()

    flat_results = flatten_dict(export_data["resultados"])

    ws_results = wb.active
    ws_results.title = "resultados_vertical"
    ws_results.append(["variable", "valor", "nombre_amigable", "modulo", "dimension", "descripcion"])

    vertical_rows = build_results_vertical_rows(flat_results)

    for row in vertical_rows:
        ws_results.append([
            row.get("variable", ""),
            row.get("valor", ""),
            row.get("nombre_amigable", ""),
            row.get("modulo", ""),
            row.get("dimension", ""),
            row.get("descripcion", "")
        ])

    style_sheet(ws_results)

    ws_matrix = wb.create_sheet("matriz_analisis")
    matrix_headers = list(flat_results.keys())
    matrix_values = [safe_value(flat_results[h]) for h in matrix_headers]
    ws_matrix.append(matrix_headers)
    ws_matrix.append(matrix_values)
    style_sheet(ws_matrix)

    add_dictionary_sheet(wb)
    add_metadata_sheet(wb, export_data["metadatos"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_rows_excel(rows, metadata, matrix_sheet_name="matriz_corpus"):
    wb = Workbook()

    ws_matrix = wb.active
    ws_matrix.title = matrix_sheet_name

    if rows:
        all_headers = []
        for row in rows:
            for key in row.keys():
                if key not in all_headers:
                    all_headers.append(key)

        ws_matrix.append(all_headers)

        for row in rows:
            ws_matrix.append([safe_value(row.get(h, "")) for h in all_headers])
    else:
        ws_matrix.append(["mensaje"])
        ws_matrix.append(["No hay registros disponibles."])

    style_sheet(ws_matrix)

    ws_vertical = wb.create_sheet("resultados_vertical")
    ws_vertical.append(["id", "variable", "valor", "nombre_amigable", "modulo", "dimension", "descripcion"])

    for idx, row in enumerate(rows, start=1):
        row_id = row.get("id") or row.get("participant_id") or row.get("submission_id") or f"registro_{idx:03d}"

        for variable, value in row.items():
            dict_info = get_dictionary_info(variable)
            ws_vertical.append([
                row_id,
                variable,
                safe_value(value),
                dict_info.get("nombre_amigable", ""),
                dict_info.get("modulo", ""),
                dict_info.get("dimension", ""),
                dict_info.get("descripcion", "")
            ])

    style_sheet(ws_vertical)

    add_dictionary_sheet(wb)
    add_metadata_sheet(wb, metadata)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_multi_analysis_excel(rows, metadata):
    return create_rows_excel(rows, metadata, matrix_sheet_name="matriz_multitexto")


def create_participant_excel():
    metadata = {
        "scriptora_version": "0.9.3",
        "timestamp_utc": datetime.utcnow().isoformat(),
        "total_respuestas": len(PARTICIPANT_SUBMISSIONS),
        "advertencia_metodologica": "Registro temporal en memoria. En Render puede perderse si el servicio se reinicia.",
        "uso": "Demo funcional para cerrar el circuito participante-investigador."
    }

    return create_rows_excel(
        rows=PARTICIPANT_SUBMISSIONS,
        metadata=metadata,
        matrix_sheet_name="respuestas_participantes"
    )


def create_corpus_excel(rows, analysis_mode):
    metadata = {
        "scriptora_version": "0.9.3",
        "timestamp_utc": datetime.utcnow().isoformat(),
        "total_textos": len(rows),
        "analysis_mode": analysis_mode,
        "archivo_generado": "scriptora_corpus_resultados_v0_9_3.xlsx",
        "formato_entrada": "Excel con columna obligatoria texto o text. Columnas opcionales: id, grupo, nivel, genero, observaciones.",
        "hoja_matriz_corpus": "Una fila por texto procesado, conservando columnas originales más variables Scriptora.",
        "advertencia_metodologica": "Resultados preliminares; requieren calibración con corpus reales, rúbricas humanas y benchmarks contextuales."
    }

    return create_rows_excel(
        rows=rows,
        metadata=metadata,
        matrix_sheet_name="matriz_corpus"
    )


# ============================================================
# INTERFAZ WEB
# ============================================================

@app.get("/", response_class=HTMLResponse)
def home():
    return """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Scriptora Suite</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f7f7fb;
            margin: 0;
            padding: 40px;
            color: #222;
        }
        .container {
            max-width: 1040px;
            margin: auto;
            background: white;
            padding: 30px;
            border-radius: 18px;
            box-shadow: 0 8px 30px rgba(0,0,0,0.08);
        }
        h1 { margin-bottom: 5px; font-size: 34px; }
        h2 { margin-top: 0; }
        .subtitle { color: #666; margin-bottom: 25px; }
        textarea, select, button, input {
            width: 100%;
            margin-top: 10px;
            margin-bottom: 15px;
            padding: 12px;
            font-size: 15px;
            border-radius: 10px;
            border: 1px solid #ccc;
            box-sizing: border-box;
        }
        textarea { height: 210px; }
        button {
            background: #111827;
            color: white;
            cursor: pointer;
            border: none;
            font-weight: bold;
        }
        button:hover { background: #374151; }
        .secondary { background: #6b7280; }
        .secondary:hover { background: #4b5563; }
        .success {
            background: #ecfdf5;
            border: 1px solid #a7f3d0;
            padding: 20px;
            border-radius: 14px;
            display: none;
            margin-top: 20px;
        }
        .result {
            margin-top: 25px;
            padding: 20px;
            background: #f0f4ff;
            border-radius: 14px;
            display: none;
        }
        .metric {
            padding: 8px 0;
            border-bottom: 1px solid #ddd;
        }
        .score-box {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 12px;
            margin-top: 10px;
        }
        .score {
            background: white;
            padding: 14px;
            border-radius: 12px;
            border: 1px solid #d8def5;
        }
        .score strong {
            display: block;
            font-size: 14px;
            color: #374151;
        }
        .score span {
            font-size: 22px;
            font-weight: bold;
        }
        .note {
            margin-top: 12px;
            font-size: 13px;
            color: #666;
            line-height: 1.4;
        }
        .pill {
            display: inline-block;
            padding: 6px 10px;
            border-radius: 999px;
            background: #111827;
            color: white;
            font-size: 13px;
            margin-bottom: 10px;
        }
        .process-panel {
            background: #eef2ff;
            border: 1px solid #c7d2fe;
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 15px;
            font-size: 14px;
        }
        .small-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
            gap: 8px;
            margin-top: 8px;
        }
        .small-box {
            background: white;
            border-radius: 10px;
            padding: 8px;
            border: 1px solid #d8def5;
        }
        .help {
            font-size: 13px;
            color: #555;
            margin-top: -8px;
            margin-bottom: 12px;
            line-height: 1.4;
        }
        .role-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }
        .role-card {
            border: 1px solid #d1d5db;
            background: #f9fafb;
            padding: 22px;
            border-radius: 16px;
        }
        .role-card h3 { margin-top: 0; }
        .hidden { display: none; }
        .topbar {
            display: flex;
            gap: 10px;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 18px;
        }
        .topbar button {
            width: auto;
            padding: 10px 16px;
            margin: 0;
        }
        .task-box {
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            padding: 18px;
            border-radius: 14px;
            margin-bottom: 18px;
        }
        .admin-box {
            background: #fff7ed;
            border: 1px solid #fed7aa;
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 15px;
        }
        .corpus-box {
            background: #f0fdf4;
            border: 1px solid #bbf7d0;
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 15px;
        }
        .config-box {
            background: #f8fafc;
            border: 1px solid #dbe3ef;
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 15px;
        }
        .format-box {
            background: white;
            border: 1px dashed #94a3b8;
            padding: 12px;
            border-radius: 10px;
            font-family: monospace;
            font-size: 13px;
            margin-top: 10px;
            line-height: 1.5;
            overflow-x: auto;
        }
    </style>
</head>
<body>
<div class="container">

    <div id="roleScreen">
        <h1>Scriptora</h1>
        <div class="subtitle">Plataforma para análisis textual, evaluación escritural y proceso de escritura · v0.9.3</div>

        <div class="role-grid">
            <div class="role-card">
                <h3>Modo participante</h3>
                <p>Para responder una tarea de escritura en una interfaz limpia, sin resultados visibles.</p>
                <button onclick="enterParticipant()">Ingresar como participante</button>
            </div>

            <div class="role-card">
                <h3>Modo investigador</h3>
                <p>Para analizar textos, evaluar escritura, procesar corpus y descargar resultados en Excel.</p>
                <button onclick="enterResearcher()">Ingresar como investigador</button>
            </div>
        </div>

        <div class="note">
            En futuras versiones se agregará ingreso con clave para participantes, investigadores y administradores.
        </div>
    </div>

    <div id="participantScreen" class="hidden">
        <div class="topbar">
            <div>
                <h1>Scriptora Participante</h1>
                <div class="subtitle">Tarea de escritura</div>
            </div>
            <button class="secondary" onclick="goHome()">Volver</button>
        </div>

        <div class="task-box">
            <h2>Tarea de escritura</h2>
            <p>Lee atentamente la consigna y escribe tu respuesta en el espacio indicado. Cuando termines, presiona “Enviar respuesta”.</p>
            <p><strong>Consigna:</strong> Escribe un texto breve en el que expreses tu opinión sobre el uso de la inteligencia artificial en educación.</p>
        </div>

        <label>ID participante opcional</label>
        <input id="participantId" placeholder="Ejemplo: sujeto_001">

        <label>Tu respuesta</label>
        <textarea id="participantText" placeholder="Escribe aquí tu respuesta..."></textarea>

        <button onclick="submitParticipantResponse()">Enviar respuesta</button>

        <div id="participantConfirmation" class="success">
            <h2>Respuesta enviada</h2>
            <p>Tu respuesta fue enviada correctamente. Muchas gracias por participar.</p>
        </div>
    </div>

    <div id="researcherScreen" class="hidden">
        <div class="topbar">
            <div>
                <h1>Scriptora Investigador</h1>
                <div class="subtitle">Análisis textual, evaluación escritural, corpus Excel y proceso de escritura</div>
            </div>
            <button class="secondary" onclick="goHome()">Volver</button>
        </div>

        <div class="admin-box">
            <strong>Registro de participantes</strong>
            <p class="note">Descarga las respuestas enviadas desde el modo participante. Registro temporal en memoria.</p>
            <button class="secondary" onclick="downloadParticipantExcel()">Descargar respuestas de participantes</button>
            <button class="secondary" onclick="clearParticipantRecords()">Limpiar registros temporales</button>
        </div>

        <div class="config-box">
            <strong>Configuración del análisis</strong>

            <label>Línea de análisis</label>
            <select id="analysisMode" onchange="updateResearcherInterface()">
                <option value="text_analysis">Scriptora T · Análisis textual descriptivo</option>
                <option value="writing_product">Scriptora W · Evaluación de producto escrito</option>
                <option value="writing_process">Scriptora W · Producto + proceso escritural</option>
            </select>

            <label>Tipo de entrada</label>
            <select id="entryMode" onchange="updateResearcherInterface()">
                <option value="single_text">Texto individual</option>
                <option value="pasted_multi">Multitexto pegado</option>
                <option value="corpus_excel">Corpus Excel</option>
                <option value="live_process">Escritura en vivo / proceso</option>
            </select>

            <div class="note" id="configHelp"></div>
        </div>

        <div id="corpusSection" class="corpus-box hidden">
            <strong>Carga de corpus Excel</strong>
            <p class="note">
                El archivo debe ser .xlsx y contener una columna obligatoria llamada <strong>texto</strong>.
                Las columnas <strong>id</strong>, <strong>grupo</strong>, <strong>nivel</strong>, <strong>genero</strong> y <strong>observaciones</strong> son recomendadas.
            </p>

            <div class="format-box">
id | texto | grupo | nivel | genero | observaciones<br>
sujeto_001 | Texto del sujeto... | grupo_A | media | argumentativo | pretest<br>
sujeto_002 | Texto del sujeto... | grupo_B | universitario | expositivo | postest
            </div>

            <p class="note">
                Si el Excel incluye <strong>nivel</strong> y <strong>genero</strong>, Scriptora usará esos valores por fila.
                Si no existen, usará valores generales. Si no incluye <strong>id</strong>, generará uno automáticamente.
            </p>

            <input type="file" id="corpusFile" accept=".xlsx">
            <button class="secondary" onclick="uploadCorpusExcel()">Procesar corpus Excel y descargar resultados</button>
        </div>

        <div id="processPanel" class="process-panel hidden">
            <strong>Registro de proceso escritural</strong>
            <div class="small-grid">
                <div class="small-box">Tiempo: <span id="timer">0</span> s</div>
                <div class="small-box">Latencia inicial: <span id="latency">0</span> s</div>
                <div class="small-box">Pausas largas: <span id="longPauses">0</span></div>
                <div class="small-box">Ediciones: <span id="edits">0</span></div>
                <div class="small-box">Inserciones: <span id="insertions">0</span></div>
                <div class="small-box">Borrados: <span id="deletions">0</span></div>
                <div class="small-box">Ajustes locales: <span id="localAdjustments">0</span></div>
                <div class="small-box">Reformulaciones: <span id="reformulations">0</span></div>
                <div class="small-box">Expansiones: <span id="expansions">0</span></div>
                <div class="small-box">Reducciones: <span id="reductions">0</span></div>
                <div class="small-box">Ajustes macro: <span id="macroAdjustments">0</span></div>
                <div class="small-box">Eventos: <span id="events">0</span></div>
            </div>
        </div>

        <div id="textSection">
            <label id="textLabel">Texto a analizar</label>
            <div id="modeHelp" class="help"></div>
            <textarea id="textInput" placeholder="Escribe aquí o pega un texto para analizar..."></textarea>
        </div>

        <div id="selectorSection">
            <label id="levelLabel">Nivel / audiencia objetivo</label>
            <select id="level">
                <option value="general">General</option>
                <option value="1_4_basico">1°–4° básico</option>
                <option value="5_8_basico">5°–8° básico</option>
                <option value="media">Enseñanza media</option>
                <option value="universitario">Universitario</option>
                <option value="adulto">Adulto general</option>
            </select>

            <label id="genreLabel">Tipo de texto</label>
            <select id="genre">
                <option value="general">General</option>
                <option value="narrativo">Narrativo</option>
                <option value="argumentativo">Argumentativo</option>
                <option value="expositivo">Expositivo</option>
                <option value="tecnico">Técnico</option>
            </select>
        </div>

        <div id="actionSection">
            <button onclick="runScriptora()">Analizar con Scriptora</button>
            <button class="secondary" onclick="downloadExcel()">Descargar Excel</button>
            <button class="secondary" onclick="resetProcessCapture()">Reiniciar captura</button>
        </div>

        <div id="result" class="result">
            <div id="moduleLabel" class="pill"></div>
            <h2>Resultado preliminar</h2>

            <h3 id="productTitle">Producto textual</h3>
            <div id="metrics"></div>
            <div id="scores"></div>

            <div id="processSectionResult" style="display:none;">
                <h3>Proceso escritural</h3>
                <div id="processMetrics"></div>
                <div id="processScores"></div>
            </div>

            <h3>Interpretación</h3>
            <p id="interpretation"></p>

            <div id="integratedSection" style="display:none;">
                <h3>Síntesis integrada</h3>
                <p id="integratedInterpretation"></p>
            </div>

            <div class="note">
                La pantalla entrega una vista preliminar. El análisis completo debe revisarse en el Excel exportado.
            </div>
        </div>
    </div>
</div>

<script>
let sessionStarted = false;
let writingStarted = false;
let processClosed = false;

let startTime = null;
let firstInputTime = null;
let lastInputTime = null;
let timerInterval = null;

let finalTotalTimeSeconds = 0;
let finalInitialLatencySeconds = 0;

let longPauseCount = 0;
let editCount = 0;
let deletionCount = 0;
let insertionCount = 0;
let localAdjustmentCount = 0;
let reformulationCount = 0;
let expansionCount = 0;
let reductionCount = 0;
let macroAdjustmentCount = 0;
let inputEventCount = 0;

let previousText = "";
let maxTextLength = 0;
let previousParagraphCount = 0;

const LONG_PAUSE_THRESHOLD_MS = 3000;
const REFORMULATION_DELTA = 25;

function enterParticipant() {
    document.getElementById("roleScreen").classList.add("hidden");
    document.getElementById("participantScreen").classList.remove("hidden");
    document.getElementById("researcherScreen").classList.add("hidden");
    document.getElementById("participantConfirmation").style.display = "none";
}

function enterResearcher() {
    document.getElementById("roleScreen").classList.add("hidden");
    document.getElementById("participantScreen").classList.add("hidden");
    document.getElementById("researcherScreen").classList.remove("hidden");
    updateResearcherInterface();
}

function goHome() {
    document.getElementById("roleScreen").classList.remove("hidden");
    document.getElementById("participantScreen").classList.add("hidden");
    document.getElementById("researcherScreen").classList.add("hidden");
}

async function submitParticipantResponse() {
    const text = document.getElementById("participantText").value;
    const participantId = document.getElementById("participantId").value;

    if (!text.trim()) {
        alert("Por favor escribe una respuesta antes de enviar.");
        return;
    }

    const response = await fetch("/api/participant/submit", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify({
            participant_id: participantId,
            text: text,
            language: "es",
            level: "general",
            genre: "argumentativo",
            task: "participant_writing_task",
            purpose: "silent_participant_submission"
        })
    });

    if (!response.ok) {
        alert("No se pudo registrar la respuesta.");
        return;
    }

    document.getElementById("participantText").value = "";
    document.getElementById("participantConfirmation").style.display = "block";
}

async function downloadParticipantExcel() {
    const response = await fetch("/api/participant/export-excel", {
        method: "GET"
    });

    if (!response.ok) {
        alert("No se pudo generar el Excel de participantes.");
        return;
    }

    const blob = await response.blob();
    downloadBlob(blob, "scriptora_respuestas_participantes_v0_9_3.xlsx");
}

async function clearParticipantRecords() {
    const ok = confirm("¿Seguro que quieres limpiar los registros temporales de participantes?");
    if (!ok) return;

    const response = await fetch("/api/participant/clear", {
        method: "POST"
    });

    if (response.ok) {
        alert("Registros temporales eliminados.");
    } else {
        alert("No se pudieron limpiar los registros.");
    }
}

async function uploadCorpusExcel() {
    const fileInput = document.getElementById("corpusFile");
    const analysisMode = document.getElementById("analysisMode").value;

    if (!fileInput.files || fileInput.files.length === 0) {
        alert("Por favor selecciona un archivo Excel .xlsx.");
        return;
    }

    const formData = new FormData();
    formData.append("file", fileInput.files[0]);
    formData.append("analysis_mode", analysisMode);

    const response = await fetch("/api/corpus/upload-excel", {
        method: "POST",
        body: formData
    });

    if (!response.ok) {
        const msg = await response.text();
        alert("No se pudo procesar el corpus Excel. Revisa que exista una columna llamada texto. Detalle: " + msg);
        return;
    }

    const blob = await response.blob();
    downloadBlob(blob, "scriptora_corpus_resultados_v0_9_3.xlsx");
}

function downloadBlob(blob, filename) {
    const url = window.URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = filename;
    document.body.appendChild(a);
    a.click();
    a.remove();
    window.URL.revokeObjectURL(url);
}

function currentAnalysisMode() {
    return document.getElementById("analysisMode").value;
}

function currentEntryMode() {
    return document.getElementById("entryMode").value;
}

function isProcessMode() {
    return currentAnalysisMode() === "writing_process" && currentEntryMode() === "live_process";
}

function isMultiMode() {
    return currentEntryMode() === "pasted_multi";
}

function isCorpusMode() {
    return currentEntryMode() === "corpus_excel";
}

function updateResearcherInterface() {
    const analysisMode = currentAnalysisMode();
    const entryModeSelect = document.getElementById("entryMode");

    if (analysisMode === "writing_process") {
        entryModeSelect.value = "live_process";
        for (const option of entryModeSelect.options) {
            option.disabled = option.value !== "live_process";
        }
    } else {
        for (const option of entryModeSelect.options) {
            option.disabled = option.value === "live_process";
        }
        if (entryModeSelect.value === "live_process") {
            entryModeSelect.value = "single_text";
        }
    }

    const entryMode = currentEntryMode();

    document.getElementById("corpusSection").classList.toggle("hidden", entryMode !== "corpus_excel");
    document.getElementById("textSection").classList.toggle("hidden", entryMode === "corpus_excel");
    document.getElementById("selectorSection").classList.toggle("hidden", entryMode === "corpus_excel");
    document.getElementById("actionSection").classList.toggle("hidden", entryMode === "corpus_excel");
    document.getElementById("processPanel").classList.toggle("hidden", !isProcessMode());

    if (!isProcessMode()) {
        resetOnlyProcessCounters();
    }

    let configHelp = "";
    let modeHelp = "";
    let textLabel = "Texto a analizar";
    let levelLabel = "Nivel / audiencia objetivo";
    let genreLabel = "Tipo de texto";

    if (analysisMode === "text_analysis") {
        configHelp = "Scriptora T describe el texto como objeto lingüístico. No entrega evaluación de desempeño escritural.";
    } else if (analysisMode === "writing_product") {
        configHelp = "Scriptora W evalúa el texto como producto escrito: organización, cohesión, elaboración, género y nivel preliminar.";
    } else {
        configHelp = "Scriptora W captura el proceso de escritura en vivo y relaciona producto final con señales de regulación escritural.";
    }

    if (entryMode === "single_text") {
        modeHelp = "Modo individual: pega o escribe un texto. Los selectores de nivel y género se aplican a este texto.";
    } else if (entryMode === "pasted_multi") {
        modeHelp = `
            Modo multitexto pegado. Usa este formato:<br>
            ### ID: sujeto_001<br>
            Texto del sujeto 001...<br><br>
            ### ID: sujeto_002<br>
            Texto del sujeto 002...<br><br>
            También puedes separar textos con una línea que contenga solo ---.<br>
            Los selectores de nivel y género se aplican como valores comunes para todos los textos.
        `;
        textLabel = "Multitexto a analizar";
        levelLabel = "Nivel común para todos los textos";
        genreLabel = "Tipo de texto común para todos los textos";
    } else if (entryMode === "corpus_excel") {
        configHelp += " En corpus Excel, si existen columnas nivel y genero, esos valores tienen prioridad por fila.";
    } else if (entryMode === "live_process") {
        modeHelp = "Modo proceso: escribe en vivo para capturar tiempo, pausas, ediciones y señales preliminares de regulación.";
        textLabel = "Texto escrito en vivo";
    }

    document.getElementById("configHelp").innerText = configHelp;
    document.getElementById("modeHelp").innerHTML = modeHelp;
    document.getElementById("textLabel").innerText = textLabel;
    document.getElementById("levelLabel").innerText = levelLabel;
    document.getElementById("genreLabel").innerText = genreLabel;

    document.getElementById("result").style.display = "none";
}

function resetOnlyProcessCounters() {
    sessionStarted = false;
    writingStarted = false;
    processClosed = false;

    startTime = null;
    firstInputTime = null;
    lastInputTime = null;

    finalTotalTimeSeconds = 0;
    finalInitialLatencySeconds = 0;

    longPauseCount = 0;
    editCount = 0;
    deletionCount = 0;
    insertionCount = 0;
    localAdjustmentCount = 0;
    reformulationCount = 0;
    expansionCount = 0;
    reductionCount = 0;
    macroAdjustmentCount = 0;
    inputEventCount = 0;

    previousText = "";
    maxTextLength = 0;
    previousParagraphCount = 0;

    if (timerInterval) {
        clearInterval(timerInterval);
        timerInterval = null;
    }

    updatePanel();
}

function startSessionIfNeeded() {
    if (!isProcessMode()) return;

    if (!sessionStarted && !processClosed) {
        sessionStarted = true;
        startTime = Date.now();
        timerInterval = setInterval(updateTimer, 1000);
    }
}

function updateTimer() {
    if (startTime && !processClosed && isProcessMode()) {
        const seconds = Math.floor((Date.now() - startTime) / 1000);
        document.getElementById("timer").innerText = seconds;
    }
}

function paragraphCount(text) {
    return text.split("\\n").filter(p => p.trim().length > 0).length;
}

function updatePanel() {
    const now = Date.now();
    const totalTimeSeconds = processClosed ? finalTotalTimeSeconds : (startTime ? Math.floor((now - startTime) / 1000) : 0);
    const latencySeconds = processClosed ? finalInitialLatencySeconds : (firstInputTime && startTime ? Math.floor((firstInputTime - startTime) / 1000) : 0);

    document.getElementById("timer").innerText = totalTimeSeconds;
    document.getElementById("latency").innerText = latencySeconds;
    document.getElementById("longPauses").innerText = longPauseCount;
    document.getElementById("edits").innerText = editCount;
    document.getElementById("insertions").innerText = insertionCount;
    document.getElementById("deletions").innerText = deletionCount;
    document.getElementById("localAdjustments").innerText = localAdjustmentCount;
    document.getElementById("reformulations").innerText = reformulationCount;
    document.getElementById("expansions").innerText = expansionCount;
    document.getElementById("reductions").innerText = reductionCount;
    document.getElementById("macroAdjustments").innerText = macroAdjustmentCount;
    document.getElementById("events").innerText = inputEventCount;
}

document.addEventListener("DOMContentLoaded", function() {
    const textarea = document.getElementById("textInput");

    textarea.addEventListener("focus", function() {
        if (isProcessMode()) {
            startSessionIfNeeded();
        }
    });

    textarea.addEventListener("paste", function() {
        if (isProcessMode()) {
            alert("Para capturar proceso, escribe en vivo. Si pegas texto, el proceso no será trazable.");
        }
    });

    textarea.addEventListener("input", function() {
        if (!isProcessMode()) return;
        if (processClosed) return;

        startSessionIfNeeded();

        const now = Date.now();
        const currentText = textarea.value;
        const currentParagraphCount = paragraphCount(currentText);

        if (!writingStarted) {
            writingStarted = true;
            firstInputTime = now;
            previousParagraphCount = currentParagraphCount;
        }

        if (lastInputTime) {
            const gap = now - lastInputTime;
            if (gap >= LONG_PAUSE_THRESHOLD_MS) {
                longPauseCount += 1;
            }
        }

        inputEventCount += 1;

        const delta = currentText.length - previousText.length;

        if (delta > 0) {
            insertionCount += delta;
            editCount += 1;
            expansionCount += 1;

            if (delta <= 15) localAdjustmentCount += 1;
            if (delta >= REFORMULATION_DELTA) reformulationCount += 1;
        }

        if (delta < 0) {
            const absDelta = Math.abs(delta);
            deletionCount += absDelta;
            editCount += 1;
            reductionCount += 1;

            if (absDelta <= 15) localAdjustmentCount += 1;
            if (absDelta >= REFORMULATION_DELTA) reformulationCount += 1;
        }

        if (currentParagraphCount !== previousParagraphCount && inputEventCount > 1) {
            macroAdjustmentCount += 1;
        }

        if (currentText.length > maxTextLength) {
            maxTextLength = currentText.length;
        }

        previousText = currentText;
        previousParagraphCount = currentParagraphCount;
        lastInputTime = now;

        updatePanel();
    });

    updateResearcherInterface();
});

function freezeProcess() {
    if (!isProcessMode()) {
        resetOnlyProcessCounters();
        return;
    }

    const now = Date.now();

    finalTotalTimeSeconds = startTime ? Math.floor((now - startTime) / 1000) : 0;
    finalInitialLatencySeconds = firstInputTime && startTime ? Math.floor((firstInputTime - startTime) / 1000) : 0;

    processClosed = true;

    if (timerInterval) {
        clearInterval(timerInterval);
        timerInterval = null;
    }

    updatePanel();
}

function resetProcessCapture() {
    resetOnlyProcessCounters();
    document.getElementById("textInput").value = "";
    document.getElementById("result").style.display = "none";
}

async function runScriptora() {
    const analysisMode = currentAnalysisMode();
    const entryMode = currentEntryMode();
    const text = document.getElementById("textInput").value;
    const level = document.getElementById("level").value;
    const genre = document.getElementById("genre").value;

    if (!text.trim()) {
        alert("Por favor escribe o pega texto antes de analizar.");
        return;
    }

    document.getElementById("result").style.display = "block";
    document.getElementById("processSectionResult").style.display = "none";
    document.getElementById("integratedSection").style.display = "none";

    if (entryMode === "pasted_multi") {
        const response = await fetch("/api/multi/analyze", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                raw_input: text,
                analysis_mode: analysisMode,
                language: "es",
                level: level,
                genre: genre,
                purpose: "multitext_analysis"
            })
        });

        const data = await response.json();

        document.getElementById("moduleLabel").innerText = data.module;
        document.getElementById("productTitle").innerText = "Resultado multitexto";
        document.getElementById("metrics").innerHTML = `
            <div class="metric"><strong>Textos procesados:</strong> ${data.total_texts}</div>
            <div class="metric"><strong>Línea de análisis:</strong> ${analysisMode}</div>
        `;

        document.getElementById("scores").innerHTML = renderMultiPreview(data.results);
        document.getElementById("interpretation").innerText = "Análisis multitexto completado. Descarga el Excel para revisar la matriz completa por texto.";
        return;
    }

    if (analysisMode === "text_analysis") {
        const response = await fetch("/api/text/analyze", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                text: text,
                language: "es",
                context: "scriptora_suite",
                level: level,
                genre: genre,
                purpose: "text_analysis"
            })
        });

        const data = await response.json();
        const metrics = data.raw_metrics;

        document.getElementById("moduleLabel").innerText = data.module;
        document.getElementById("productTitle").innerText = "Análisis textual descriptivo";

        document.getElementById("metrics").innerHTML = `
            <div class="metric"><strong>Palabras:</strong> ${metrics.word_count}</div>
            <div class="metric"><strong>Caracteres:</strong> ${metrics.char_count}</div>
            <div class="metric"><strong>Oraciones:</strong> ${metrics.sentence_count}</div>
            <div class="metric"><strong>Palabras únicas:</strong> ${metrics.unique_words}</div>
            <div class="metric"><strong>Longitud oracional promedio:</strong> ${metrics.avg_sentence_length}</div>
            <div class="metric"><strong>TTR:</strong> ${metrics.type_token_ratio}</div>
            <div class="metric"><strong>Densidad léxica estimada:</strong> ${metrics.lexical_density_proxy}</div>
        `;

        document.getElementById("scores").innerHTML = "";
        document.getElementById("interpretation").innerText = data.interpretation;
        return;
    }

    if (analysisMode === "writing_product") {
        const response = await fetch("/api/write/evaluate", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                text: text,
                language: "es",
                level: level,
                genre: genre,
                task: "open_writing_task",
                purpose: "writing_product_evaluation"
            })
        });

        const data = await response.json();
        document.getElementById("moduleLabel").innerText = data.module;
        renderWritingProduct(data.writing_metrics);
        return;
    }

    if (analysisMode === "writing_process") {
        if (!processClosed) {
            freezeProcess();
        }

        const response = await fetch("/api/write/process-evaluate", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                text: text,
                language: "es",
                level: level,
                genre: genre,
                task: "open_writing_task",
                purpose: "process_writing_evaluation",
                writing_mode: "live",
                total_time_seconds: finalTotalTimeSeconds,
                initial_latency_seconds: finalInitialLatencySeconds,
                pause_count: longPauseCount,
                long_pause_count: longPauseCount,
                edit_count: editCount,
                deletion_count: deletionCount,
                insertion_count: insertionCount,
                local_adjustment_count: localAdjustmentCount,
                reformulation_count: reformulationCount,
                expansion_count: expansionCount,
                reduction_count: reductionCount,
                macro_adjustment_count: macroAdjustmentCount,
                max_text_length: maxTextLength || text.length,
                final_text_length: text.length,
                input_event_count: inputEventCount
            })
        });

        const data = await response.json();

        document.getElementById("moduleLabel").innerText = data.module;

        const product = data.product_metrics;
        const process = data.process_metrics;

        renderWritingProduct(product);

        document.getElementById("processSectionResult").style.display = "block";

        document.getElementById("processMetrics").innerHTML = `
            <div class="metric"><strong>Modo de ingreso:</strong> ${process.writing_mode}</div>
            <div class="metric"><strong>Tiempo total:</strong> ${process.total_time_seconds} s</div>
            <div class="metric"><strong>Latencia inicial:</strong> ${process.initial_latency_seconds} s</div>
            <div class="metric"><strong>Pausas largas:</strong> ${process.long_pause_count}</div>
            <div class="metric"><strong>Ediciones:</strong> ${process.edit_count}</div>
            <div class="metric"><strong>Inserciones:</strong> ${process.insertion_count}</div>
            <div class="metric"><strong>Borrados:</strong> ${process.deletion_count}</div>
            <div class="metric"><strong>Ajustes locales:</strong> ${process.local_adjustment_count}</div>
            <div class="metric"><strong>Reformulaciones:</strong> ${process.reformulation_count}</div>
            <div class="metric"><strong>Expansiones:</strong> ${process.expansion_count}</div>
            <div class="metric"><strong>Reducciones:</strong> ${process.reduction_count}</div>
            <div class="metric"><strong>Ajustes macrotextuales:</strong> ${process.macro_adjustment_count}</div>
            <div class="metric"><strong>Eventos de escritura:</strong> ${process.input_event_count}</div>
            <div class="metric"><strong>Palabras por minuto:</strong> ${process.words_per_minute}</div>
            <div class="metric"><strong>Estabilidad final:</strong> ${process.final_stability_ratio}</div>
        `;

        document.getElementById("processScores").innerHTML = `
            <h3>Puntajes preliminares del proceso</h3>
            <div class="score-box">
                <div class="score"><strong>Planificación</strong><span>${process.planning_score}</span></div>
                <div class="score"><strong>Monitoreo</strong><span>${process.monitoring_score}</span></div>
                <div class="score"><strong>Revisión</strong><span>${process.revision_score}</span></div>
                <div class="score"><strong>Reformulación</strong><span>${process.reformulation_score}</span></div>
                <div class="score"><strong>Fluidez</strong><span>${process.fluency_score}</span></div>
                <div class="score"><strong>Recursividad</strong><span>${process.recursivity_score}</span></div>
                <div class="score"><strong>Regulación global</strong><span>${process.process_regulation_score}</span></div>
                <div class="score"><strong>Nivel proceso</strong><span>${process.process_regulation_label}</span></div>
            </div>
        `;

        document.getElementById("interpretation").innerText = process.interpretation;
        document.getElementById("integratedSection").style.display = "block";
        document.getElementById("integratedInterpretation").innerText = data.integrated_interpretation;
    }
}

function renderWritingProduct(metrics) {
    const scores = metrics.scores;

    document.getElementById("productTitle").innerText = "Producto escrito";

    document.getElementById("metrics").innerHTML = `
        <div class="metric"><strong>Palabras:</strong> ${metrics.word_count}</div>
        <div class="metric"><strong>Oraciones:</strong> ${metrics.sentence_count}</div>
        <div class="metric"><strong>Párrafos:</strong> ${metrics.paragraph_count}</div>
        <div class="metric"><strong>Longitud oracional promedio:</strong> ${metrics.avg_sentence_length}</div>
        <div class="metric"><strong>TTR:</strong> ${metrics.type_token_ratio}</div>
        <div class="metric"><strong>Densidad léxica:</strong> ${metrics.lexical_density_proxy}</div>
        <div class="metric"><strong>Conectores detectados:</strong> ${metrics.connector_count}</div>
        <div class="metric"><strong>Conectores:</strong> ${metrics.connectors_found.join(", ") || "No detectados"}</div>
        <div class="metric"><strong>Marcas de puntuación:</strong> ${metrics.punctuation_count}</div>
        <div class="metric"><strong>Cierre textual:</strong> ${metrics.closure_present ? "Detectado" : "No detectado"}</div>
    `;

    document.getElementById("scores").innerHTML = `
        <h3>Puntajes preliminares del producto</h3>
        <div class="score-box">
            <div class="score"><strong>Extensión</strong><span>${scores.extension}</span></div>
            <div class="score"><strong>Organización</strong><span>${scores.organization}</span></div>
            <div class="score"><strong>Cohesión</strong><span>${scores.cohesion}</span></div>
            <div class="score"><strong>Sintaxis</strong><span>${scores.syntax_control}</span></div>
            <div class="score"><strong>Léxico</strong><span>${scores.lexical_variety}</span></div>
            <div class="score"><strong>Densidad informativa</strong><span>${scores.informational_density}</span></div>
            <div class="score"><strong>Elaboración</strong><span>${scores.idea_elaboration}</span></div>
            <div class="score"><strong>Coherencia global</strong><span>${scores.global_coherence}</span></div>
            <div class="score"><strong>Puntuación</strong><span>${scores.punctuation_control}</span></div>
            <div class="score"><strong>Adecuación género</strong><span>${scores.genre_adequacy}</span></div>
            <div class="score"><strong>Cierre</strong><span>${scores.textual_closure}</span></div>
            <div class="score"><strong>Puntaje global</strong><span>${scores.global_writing_score}</span></div>
            <div class="score"><strong>Nivel</strong><span>${metrics.level_label}</span></div>
        </div>
    `;

    document.getElementById("interpretation").innerText = metrics.interpretation;
}

function renderMultiPreview(results) {
    if (!results || results.length === 0) {
        return "<p>No se detectaron textos.</p>";
    }

    let html = "<h3>Vista preliminar</h3>";
    html += "<div class='score-box'>";

    results.slice(0, 8).forEach((item) => {
        const score = item.scores_global_writing_score || item.word_count || "";
        html += `
            <div class="score">
                <strong>${item.id}</strong>
                <span>${score}</span>
            </div>
        `;
    });

    html += "</div>";

    if (results.length > 8) {
        html += `<p class="note">Se muestran 8 de ${results.length} textos. Descarga el Excel para ver todo.</p>`;
    }

    return html;
}

async function downloadExcel() {
    const analysisMode = currentAnalysisMode();
    const entryMode = currentEntryMode();
    const text = document.getElementById("textInput").value;
    const level = document.getElementById("level").value;
    const genre = document.getElementById("genre").value;

    if (!text.trim()) {
        alert("Por favor escribe o pega texto antes de descargar el Excel.");
        return;
    }

    if (entryMode === "pasted_multi") {
        const response = await fetch("/api/export/multi-excel", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                raw_input: text,
                analysis_mode: analysisMode,
                language: "es",
                level: level,
                genre: genre,
                purpose: "multitext_excel_export"
            })
        });

        if (!response.ok) {
            alert("No se pudo generar el Excel multitexto.");
            return;
        }

        const blob = await response.blob();
        downloadBlob(blob, "scriptora_multitexto_v0_9_3.xlsx");
        return;
    }

    if (analysisMode === "writing_process" && !processClosed) {
        freezeProcess();
    }

    const payload = {
        analysis_mode: analysisMode,
        entry_mode: entryMode,
        text: text,
        language: "es",
        level: level,
        genre: genre,
        task: "open_writing_task",
        purpose: "scriptora_excel_export",
        writing_mode: isProcessMode() ? "live" : "pasted",
        total_time_seconds: finalTotalTimeSeconds,
        initial_latency_seconds: finalInitialLatencySeconds,
        pause_count: longPauseCount,
        long_pause_count: longPauseCount,
        edit_count: editCount,
        deletion_count: deletionCount,
        insertion_count: insertionCount,
        local_adjustment_count: localAdjustmentCount,
        reformulation_count: reformulationCount,
        expansion_count: expansionCount,
        reduction_count: reductionCount,
        macro_adjustment_count: macroAdjustmentCount,
        max_text_length: maxTextLength || text.length,
        final_text_length: text.length,
        input_event_count: inputEventCount
    };

    const response = await fetch("/api/export/excel", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify(payload)
    });

    if (!response.ok) {
        alert("No se pudo generar el Excel.");
        return;
    }

    const blob = await response.blob();
    downloadBlob(blob, "scriptora_resultados_v0_9_3.xlsx");
}
</script>
</body>
</html>
"""


# ============================================================
# ENDPOINTS
# ============================================================

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "service": "Scriptora Suite",
        "version": "0.9.3",
        "timestamp": datetime.utcnow().isoformat(),
        "participant_submissions": len(PARTICIPANT_SUBMISSIONS)
    }


@app.post("/api/text/analyze")
def analyze_text(request: TextAnalysisRequest):
    metrics = analyze_text_core(request.text)

    return {
        "module": "Scriptora T · Text Analysis",
        "version": "0.9.3",
        "language": request.language,
        "context": request.context,
        "level": request.level,
        "genre": request.genre,
        "purpose": request.purpose,
        "raw_metrics": {
            "word_count": metrics["word_count"],
            "char_count": metrics["char_count"],
            "sentence_count": metrics["sentence_count"],
            "unique_words": metrics["unique_words"],
            "avg_sentence_length": metrics["avg_sentence_length"],
            "type_token_ratio": metrics["type_token_ratio"],
            "lexical_density_proxy": metrics["lexical_density_proxy"]
        },
        "interpretation": metrics["interpretation_text"]
    }


@app.post("/api/write/evaluate")
def evaluate_writing(request: WritingEvaluationRequest):
    writing_metrics = score_writing_product(
        text=request.text,
        genre=request.genre or "general"
    )

    return {
        "module": "Scriptora W · Writing Product Evaluation",
        "version": "0.9.3",
        "language": request.language,
        "level": request.level,
        "genre": request.genre,
        "task": request.task,
        "purpose": request.purpose,
        "writing_metrics": writing_metrics
    }


@app.post("/api/write/process-evaluate")
def evaluate_writing_process(request: WritingProcessRequest):
    product_metrics = score_writing_product(
        text=request.text,
        genre=request.genre or "general"
    )

    process_metrics = score_writing_process(request)

    integrated = integrated_writing_interpretation(
        product_metrics=product_metrics,
        process_metrics=process_metrics
    )

    return {
        "module": "Scriptora W · Product + Process Evaluation",
        "version": "0.9.3",
        "language": request.language,
        "level": request.level,
        "genre": request.genre,
        "task": request.task,
        "purpose": request.purpose,
        "product_metrics": product_metrics,
        "process_metrics": process_metrics,
        "integrated_interpretation": integrated
    }


@app.post("/api/multi/analyze")
def multi_analyze(request: MultiTextRequest):
    results = analyze_multitext(request)

    module_label = "Scriptora T · Multitexto" if request.analysis_mode == "text_analysis" else "Scriptora W · Producto multitexto"

    return {
        "module": module_label,
        "version": "0.9.3",
        "total_texts": len(results),
        "language": request.language,
        "level": request.level,
        "genre": request.genre,
        "analysis_mode": request.analysis_mode,
        "results": results
    }


@app.post("/api/participant/submit")
def participant_submit(request: ParticipantSubmissionRequest):
    record = create_participant_record(request)

    return {
        "status": "ok",
        "message": "Respuesta registrada correctamente.",
        "submission_id": record["submission_id"],
        "participant_id": record["participant_id"],
        "total_submissions": len(PARTICIPANT_SUBMISSIONS)
    }


@app.get("/api/participant/list")
def participant_list():
    return {
        "status": "ok",
        "total_submissions": len(PARTICIPANT_SUBMISSIONS),
        "records": PARTICIPANT_SUBMISSIONS
    }


@app.get("/api/participant/export-excel")
def participant_export_excel():
    output = create_participant_excel()
    filename = f"scriptora_respuestas_participantes_v0_9_3_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.post("/api/participant/clear")
def participant_clear():
    PARTICIPANT_SUBMISSIONS.clear()

    return {
        "status": "ok",
        "message": "Registros temporales eliminados.",
        "total_submissions": len(PARTICIPANT_SUBMISSIONS)
    }


@app.post("/api/corpus/upload-excel")
async def corpus_upload_excel(
    file: UploadFile = File(...),
    analysis_mode: str = Form("writing_product")
):
    if not file.filename.lower().endswith(".xlsx"):
        return HTMLResponse(
            content="El archivo debe tener formato .xlsx",
            status_code=400
        )

    try:
        file_bytes = await file.read()
        rows = process_corpus_excel(file_bytes, analysis_mode=analysis_mode)
        output = create_corpus_excel(rows, analysis_mode=analysis_mode)

        filename = f"scriptora_corpus_resultados_v0_9_3_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        return HTMLResponse(
            content=f"Error al procesar el corpus Excel: {str(e)}",
            status_code=400
        )


@app.post("/api/export/excel")
def export_excel(request: ExcelExportRequest):
    analysis_id = str(uuid.uuid4())
    timestamp = datetime.utcnow().isoformat()

    text = request.text.strip()
    analysis_mode = request.analysis_mode
    entry_mode = request.entry_mode

    resultados = {
        "analysis_id": analysis_id,
        "timestamp_utc": timestamp,
        "scriptora_version": "0.9.3",
        "analysis_mode": analysis_mode,
        "entry_mode": entry_mode,
        "language": request.language,
        "level": request.level,
        "genre": request.genre,
        "task": request.task,
        "purpose": request.purpose,
        "writing_mode": request.writing_mode,
        "texto_original": text
    }

    if analysis_mode == "text_analysis":
        metrics = analyze_text_core(text)

        resultados.update({
            "module": "Scriptora T",
            **metrics
        })

    elif analysis_mode == "writing_product":
        product_metrics = score_writing_product(
            text=text,
            genre=request.genre or "general"
        )

        resultados.update({
            "module": "Scriptora W Producto",
            "interpretation_product": product_metrics.get("interpretation", "")
        })

        resultados.update(flatten_dict(product_metrics))

    else:
        product_metrics = score_writing_product(
            text=text,
            genre=request.genre or "general"
        )

        process_metrics = score_writing_process(request)

        integrated = integrated_writing_interpretation(
            product_metrics=product_metrics,
            process_metrics=process_metrics
        )

        resultados.update({
            "module": "Scriptora W Producto + Proceso",
            "interpretation_product": product_metrics.get("interpretation", ""),
            "interpretation_process": process_metrics.get("interpretation", ""),
            "integrated_interpretation": integrated
        })

        resultados.update(flatten_dict(product_metrics, "product"))
        resultados.update(flatten_dict(process_metrics, "process"))

    metadatos = {
        "analysis_id": analysis_id,
        "timestamp_utc": timestamp,
        "scriptora_version": "0.9.3",
        "analysis_mode": analysis_mode,
        "entry_mode": entry_mode,
        "archivo_generado": "scriptora_resultados_v0_9_3.xlsx",
        "hoja_resultados_vertical": "Resultados en formato vertical legible.",
        "hoja_matriz_analisis": "Resultados en formato horizontal para análisis estadístico.",
        "hoja_diccionario_variables": "Definiciones operativas de las variables incluidas.",
        "hoja_metadatos": "Información general del análisis y versión.",
        "advertencia_metodologica": "Resultados preliminares; deben calibrarse con datos reales, rúbricas humanas y benchmarks contextuales.",
        "trazabilidad_proceso": "La interpretación del proceso solo es válida cuando el texto fue escrito en vivo y no pegado."
    }

    excel_data = {
        "resultados": resultados,
        "metadatos": metadatos
    }

    output = create_single_analysis_excel(excel_data)

    filename = f"scriptora_resultados_v0_9_3_{analysis_id[:8]}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.post("/api/export/multi-excel")
def export_multi_excel(request: MultiExcelExportRequest):
    analysis_id = str(uuid.uuid4())
    timestamp = datetime.utcnow().isoformat()

    rows = analyze_multitext(request)

    metadata = {
        "analysis_id": analysis_id,
        "timestamp_utc": timestamp,
        "scriptora_version": "0.9.3",
        "archivo_generado": "scriptora_multitexto_v0_9_3.xlsx",
        "analysis_mode": request.analysis_mode,
        "language": request.language,
        "level": request.level,
        "genre": request.genre,
        "total_texts": len(rows),
        "formato_entrada": "Bloques con ### ID: o separación mediante línea ---.",
        "hoja_matriz_multitexto": "Una fila por texto procesado.",
        "hoja_resultados_vertical": "Resultados en formato vertical por ID y variable.",
        "hoja_diccionario_variables": "Definiciones operativas de variables.",
        "advertencia_metodologica": "Resultados preliminares; requieren calibración con corpus reales, rúbricas humanas y benchmarks contextuales."
    }

    output = create_multi_analysis_excel(rows, metadata)

    filename = f"scriptora_multitexto_v0_9_3_{analysis_id[:8]}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.get("/api/benchmarks")
def list_benchmarks():
    return {
        "benchmarks": [
            {
                "benchmark_id": "ES_TEXT_GENERAL_V1",
                "module": "Scriptora T",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_WRITE_PRODUCT_V1",
                "module": "Scriptora W · Product",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_WRITE_PROCESS_REGULATION_V1",
                "module": "Scriptora W · Process",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_TEXT_MULTI_V1",
                "module": "Scriptora T · Multitexto",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_WRITE_PRODUCT_MULTI_V1",
                "module": "Scriptora W · Producto multitexto",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_PARTICIPANT_REGISTRY_V1",
                "module": "Scriptora · Registro participantes",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_CORPUS_EXCEL_UPLOAD_V2",
                "module": "Scriptora · Corpus Excel",
                "language": "es",
                "status": "prototype"
            },
            {
                "benchmark_id": "ES_RESEARCHER_INTERFACE_CLEANUP_V1",
                "module": "Scriptora · Interfaz investigador",
                "language": "es",
                "status": "prototype"
            }
        ]
    }
